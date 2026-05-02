#!/usr/bin/env python3
"""
excel_parser.py — Parse ActivoBank Excel export files

ActivoBank format:
  Rows 1–7: header metadata (account number, currency, date range, blank rows, column headers)
  Row 8 onwards: transaction data
  Columns: Data Lanc. | Data Valor | Descrição | Valor | Saldo

Usage (CLI):
  python3 excel_parser.py <path_to_excel.xlsx> [--json] [--account personal|joint]
"""

from __future__ import annotations

import io
import json
import sys
from dataclasses import dataclass, asdict
from datetime import date, datetime
from typing import Optional

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class Transaction:
    posting_date: date
    value_date: date
    description: str
    amount: float
    balance: float
    account: str = "unknown"   # "personal" or "joint", set by caller

    @property
    def is_expense(self) -> bool:
        return self.amount < 0

    def to_dict(self) -> dict:
        d = asdict(self)
        d["posting_date"] = self.posting_date.isoformat()
        d["value_date"] = self.value_date.isoformat()
        d["is_expense"] = self.is_expense
        return d


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------

_HEADER_ROWS = 7   # ActivoBank: data starts at row 8


def _parse_date(value) -> Optional[date]:
    """Convert various date representations from openpyxl to a date object."""
    if value is None:
        return None
    if isinstance(value, (datetime,)):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def parse_activobank_export(
    data: bytes,
    account: str = "unknown",
) -> list[Transaction]:
    """
    Parse an ActivoBank Excel export (bytes) and return a list of Transaction objects.

    Args:
        data:    Raw Excel file bytes (e.g., from file.read() or Telegram attachment).
        account: Label for the account ("personal" or "joint").

    Returns:
        List of Transaction objects in file order, skipping blank/header rows.
    """
    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active

    transactions: list[Transaction] = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # Skip header rows (0-indexed internally, header rows = first 7)
        if i < _HEADER_ROWS:
            continue

        if len(row) < 5:
            continue

        posting_raw, value_raw, description, amount, balance = row[:5]

        # Skip blank rows
        if posting_raw is None and description is None and amount is None:
            continue

        posting_date = _parse_date(posting_raw)
        value_date = _parse_date(value_raw)

        if posting_date is None or value_date is None:
            continue
        if description is None:
            continue
        if amount is None:
            continue

        transactions.append(Transaction(
            posting_date=posting_date,
            value_date=value_date,
            description=str(description).strip(),
            amount=round(float(amount), 2),
            balance=round(float(balance), 2) if balance is not None else 0.0,
            account=account,
        ))

    wb.close()
    return transactions


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(description="Parse an ActivoBank Excel export")
    parser.add_argument("file", help="Path to the Excel file")
    parser.add_argument(
        "--account",
        choices=["personal", "joint"],
        default="unknown",
        help="Which account this export belongs to",
    )
    parser.add_argument("--json", action="store_true", help="Output JSON array")
    args = parser.parse_args()

    with open(args.file, "rb") as f:
        data = f.read()

    transactions = parse_activobank_export(data, account=args.account)

    if args.json:
        print(json.dumps([t.to_dict() for t in transactions], ensure_ascii=False, indent=2))
    else:
        for t in transactions:
            sign = "-" if t.is_expense else "+"
            print(f"{t.posting_date}  {sign}{abs(t.amount):>10.2f} €  {t.description[:60]}")
        print(f"\nTotal: {len(transactions)} transactions")


if __name__ == "__main__":
    main()
