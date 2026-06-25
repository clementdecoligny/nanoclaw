#!/opt/wpenv/bin/python3
"""
Categorize transactions using a two-pass hybrid strategy:
  Pass 1 — exact lookup against historical data (fast, free, deterministic)
  Pass 2 — emit uncategorized rows as JSON for Edmond to classify via Claude

Key rules:
- Last occurrence in history wins (most recent categorization takes priority)
- Conflicting historical categories for the same description are flagged
- Credits that don't fit a debit category → INCOME
- Target month filtering by Data Valor before categorization
"""

import sys
import json
import argparse
from pathlib import Path
from collections import defaultdict
import openpyxl
from utils import normalize, find_category_header_row


TAXONOMY = {
    "BABYSITTING": ["BABYSITTING"],
    "EAT OUT": ["BACKERY", "PADARIA", "QUIOSQUE", "RESTAURANT", "UBER EATS"],
    "EDUCATION": ["ESCOLA"],
    "EMPREGADA": ["EMPREGADA", "SEGURANCA SOCIAL"],
    "GROCERIES": ["SUPER", "SUPER ONLINE"],
    "HEALTH": ["COUCHES", "HOSPITAL", "MEDIS", "MEDIS REIMBURSMENT", "PHARMACY"],
    "HOLIDAY": ["ALPS JULY 2026", "BIKE", "FOOD", "HOTELS", "MISC", "TRANSPORT"],
    "HOUSE": ["DONA AJUDA", "EDP", "EPAL AGUA", "FURNITURE", "MISC", "NOS", "NOS INTERNET", "SPOTIFY"],
    "KIDS": ["BABYSITTING", "LEISURE", "MISC", "PISCINA", "ROPA", "ROPA VINTED"],
    "MOBILITY": ["BOLT", "COOLTRA", "PUBLIC TRANSPORT", "TRANSPORT", "UBER"],
    "RENT": ["RENT"],
    "VOITURE": ["ESSENCE", "MECHANIC", "VIA VERDE"],
    "BIKE": ["GEAR", "MAINTENANCE", "RENTAL"],
    "CASH": ["CASH"],
    "LEISURE": [],
    "FURNITURE": ["FURNITURE"],
    "GIFT": ["GIFT"],
    "INCOME": ["SALARY", "TRANSFER", "REIMBURSEMENT", "OTHER"],
    "MISC": ["MISC"],
    "UNCLASSIFIED": ["UNCLASSIFIED"],
}

# Categories that represent money coming in rather than spending
INCOME_CATEGORIES = {"INCOME"}


def _parse_history_xlsx(xlsx_path):
    """Extract {normalized_description: [(category, sub_category, file_mtime)]} from one history file."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    header_row_idx = find_category_header_row(ws)
    if not header_row_idx:
        return {}

    col = {}
    for i, cell in enumerate(ws[header_row_idx]):
        if not cell.value:
            continue
        h = str(cell.value).strip().upper()
        if h in ("DESCRIPTION", "DESCRIÇÃO"):
            col["desc"] = i
        elif h == "CATEGORY":
            col["cat"] = i
        elif h == "SUB-CATEGORY":
            col["sub"] = i

    if "desc" not in col or "cat" not in col:
        return {}

    entries = {}
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        desc = row[col["desc"]] if len(row) > col["desc"] else None
        cat = row[col["cat"]] if len(row) > col["cat"] else None
        sub = row[col["sub"]] if "sub" in col and len(row) > col["sub"] else None
        if desc and cat:
            key = normalize(str(desc))
            entries[key] = {
                "category": str(cat).strip().upper(),
                "sub_category": str(sub).strip().upper() if sub else "",
                "raw_description": str(desc).strip(),
            }
    return entries


def load_history(history_dir):
    """Build lookup and conflict map from historical files.

    Last occurrence wins — files are sorted by name (YYYY-MM prefix),
    so the most recent month's categorization overrides older ones.

    Returns:
        lookup: {normalized_desc: {category, sub_category, raw_description}}
        conflicts: {normalized_desc: [(cat_a, file_a), (cat_b, file_b)]} — only entries with >1 distinct category
    """
    history_path = Path(history_dir)
    lookup = {}
    # Track all seen categories per description to detect conflicts
    seen = defaultdict(dict)  # key → {category: filename}

    if not history_path.exists():
        return lookup, {}

    for xlsx in sorted(history_path.glob("*.xlsx")):
        try:
            entries = _parse_history_xlsx(xlsx)
            for key, val in entries.items():
                seen[key][val["category"]] = xlsx.name
                # Last occurrence wins — overwrite unconditionally
                lookup[key] = val
        except Exception as e:
            print(f"Warning: could not read {xlsx}: {e}", file=sys.stderr)

    # Overlay learned_categories.json (highest priority — Claude-confirmed)
    learned_path = history_path / "learned_categories.json"
    if learned_path.exists():
        try:
            with open(learned_path, encoding="utf-8") as f:
                learned = json.load(f)
            for raw_key, val in learned.items():
                nkey = normalize(raw_key)
                lookup[nkey] = val
                # Learned entries resolve conflicts — remove from seen
                seen.pop(nkey, None)
        except Exception as e:
            print(f"Warning: could not read learned_categories.json: {e}", file=sys.stderr)

    # Build conflict map: descriptions that had >1 distinct category across files
    conflicts = {
        key: list(cat_files.items())
        for key, cat_files in seen.items()
        if len(cat_files) > 1
    }

    return lookup, conflicts


def filter_to_month(transactions, year, month):
    """Keep only transactions whose Data Valor falls in the given year/month."""
    prefix = f"{year:04d}-{month:02d}"
    return [t for t in transactions if (t.get("value_date") or "").startswith(prefix)]


def categorize(transactions, lookup, conflicts):
    """Apply lookup. Returns (transactions, unknown_indices, flagged_conflicts).

    flagged_conflicts: list of {description, options} for transactions whose
    description has conflicting historical categories.
    """
    unknown = []
    flagged_conflicts = []

    for i, t in enumerate(transactions):
        # Already forced (blank description)
        if t.get("confidence") == "forced":
            continue

        key = normalize(t["description"])

        if key in conflicts:
            # Multiple historical categories — flag for user resolution
            flagged_conflicts.append({
                "index": i,
                "description": t["description"],
                "amount": t["amount"],
                "options": conflicts[key],
            })
            unknown.append(i)
        elif key in lookup:
            t["category"] = lookup[key]["category"]
            t["sub_category"] = lookup[key]["sub_category"]
            t["confidence"] = "exact"
        else:
            unknown.append(i)

    return transactions, unknown, flagged_conflicts


def build_claude_prompt(transactions, unknown_indices, lookup, sample_size=25):
    """Build prompt payload for Claude to classify unknown transactions."""
    examples = []
    seen_cats = set()
    for key, val in lookup.items():
        cat = val["category"]
        if cat not in seen_cats and len(examples) < sample_size:
            examples.append({"description": key, "category": cat, "sub_category": val["sub_category"]})
            seen_cats.add(cat)

    rows_to_classify = [
        {"index": i, "description": transactions[i]["description"], "amount": transactions[i]["amount"]}
        for i in unknown_indices
    ]

    return {"taxonomy": TAXONOMY, "examples": examples, "rows": rows_to_classify}


def save_learned(history_dir, new_categorizations):
    """Persist new Claude-assigned categorizations to learned_categories.json."""
    learned_path = Path(history_dir) / "learned_categories.json"
    existing = {}
    if learned_path.exists():
        try:
            with open(learned_path, encoding="utf-8") as f:
                existing = json.load(f)
        except Exception:
            pass
    existing.update(new_categorizations)
    with open(learned_path, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("transactions_json")
    parser.add_argument("--history", required=True, help="Directory with historical .xlsx files")
    parser.add_argument("--output", required=True, help="Output JSON path")
    parser.add_argument("--year", type=int, required=True, help="Target year (e.g. 2026)")
    parser.add_argument("--month", type=int, required=True, help="Target month (1-12)")
    parser.add_argument("--apply-claude", default=None,
                        help="JSON file with Claude's categorization results to apply and persist")
    args = parser.parse_args()

    with open(args.transactions_json, encoding="utf-8") as f:
        data = json.load(f)

    # Accept both raw list and parser output dict
    transactions = data if isinstance(data, list) else data.get("transactions", [])

    # Filter to target month by Data Valor
    transactions = filter_to_month(transactions, args.year, args.month)

    lookup, conflicts = load_history(args.history)
    transactions, unknown_indices, flagged_conflicts = categorize(transactions, lookup, conflicts)

    if args.apply_claude and Path(args.apply_claude).exists():
        with open(args.apply_claude, encoding="utf-8") as f:
            claude_results = json.load(f)
        new_learned = {}
        for r in claude_results:
            idx = r["index"]
            if 0 <= idx < len(transactions):
                conf = r.get("confidence", "medium")  # high | medium | low
                transactions[idx]["category"] = r.get("category", "UNCLASSIFIED")
                transactions[idx]["sub_category"] = r.get("sub_category", "")
                transactions[idx]["confidence"] = conf
                # Only persist high-confidence classifications to the lookup
                if conf == "high" and transactions[idx]["category"] != "UNCLASSIFIED":
                    new_learned[transactions[idx]["description"]] = {
                        "category": transactions[idx]["category"],
                        "sub_category": transactions[idx]["sub_category"],
                    }
        if new_learned:
            save_learned(args.history, new_learned)
        unknown_indices = [i for i in unknown_indices if transactions[i].get("confidence") != "claude"]
        flagged_conflicts = [f for f in flagged_conflicts if f["index"] in unknown_indices]

    exact_count = sum(1 for t in transactions if t.get("confidence") == "exact")

    result = {
        "transactions": transactions,
        "unknown_count": len(unknown_indices),
        "exact_count": exact_count,
        "conflict_count": len(flagged_conflicts),
    }

    if unknown_indices:
        result["claude_prompt"] = build_claude_prompt(transactions, unknown_indices, lookup)
        result["unknown_indices"] = unknown_indices

    if flagged_conflicts:
        result["conflicts"] = flagged_conflicts

    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(
        f"Categorized {len(transactions)} transactions: "
        f"{exact_count} exact, {len(unknown_indices)} need Claude, "
        f"{len(flagged_conflicts)} conflicts.",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
