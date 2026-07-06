#!/usr/bin/env python3
"""
Migrate all transaction and income data into finance.db (SQLite).

Sources — transactions:
  - /workspace/extra/historical/historical_data_commun.xlsx  (COMUN DATA sheet, Jan 2025 - Mar 2026)
  - /workspace/extra/historical/historical_data_perso.xlsx   (PERSO DATA sheet, Feb 2025 - Mar 2026)
  - /workspace/agent/finance/historical/YYYY-MM-joint.json   (Apr 2026+)
  - /workspace/agent/finance/historical/YYYY-MM-personal.json (Apr 2026+)

Sources — income:
  - /workspace/extra/historical/historical_data_income.xlsx  (INCOME sheet, Jan 2025 - Mar 2026)
  - /workspace/agent/finance/historical/YYYY-MM-income.json  (Apr 2026+)
"""

import sqlite3
import json
import glob
import os
import sys
import openpyxl
from datetime import datetime

DB_PATH = "/workspace/agent/finance/finance.db"
HIST_DIR = "/workspace/agent/finance/historical"
HIST_COMMUN = "/workspace/extra/historical/historical_data_commun.xlsx"
HIST_PERSO = "/workspace/extra/historical/historical_data_perso.xlsx"


def connect():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def create_schema(conn):
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS transactions (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            date        TEXT    NOT NULL,
            description TEXT,
            amount      REAL    NOT NULL,
            account     TEXT    NOT NULL,
            category    TEXT,
            subcategory TEXT,
            year        INTEGER NOT NULL,
            month       INTEGER NOT NULL,
            source      TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_date        ON transactions(date);
        CREATE INDEX IF NOT EXISTS idx_account     ON transactions(account);
        CREATE INDEX IF NOT EXISTS idx_category    ON transactions(category);
        CREATE INDEX IF NOT EXISTS idx_year_month  ON transactions(year, month);
    """)
    conn.commit()


def insert_batch(conn, rows):
    conn.executemany(
        "INSERT INTO transactions (date, description, amount, account, category, subcategory, year, month, source) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        rows
    )
    conn.commit()


def import_xlsx_commun(conn):
    wb = openpyxl.load_workbook(HIST_COMMUN, data_only=True)
    ws = wb["COMUN DATA"]
    rows = []
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = row[0]
        if not d or not hasattr(d, 'year'):
            skipped += 1
            continue
        # Import all entries — xlsx has escola manual entries for Apr 2026 (Coverflex)
        # that are NOT in our JSONs, so include them all
        date_str = d.strftime("%Y-%m-%d")
        desc = row[1] or ""
        amount = row[2]
        if amount is None:
            skipped += 1
            continue
        category = row[3] or ""
        subcategory = row[4] or ""
        rows.append((date_str, desc, float(amount), "joint",
                     category, subcategory, d.year, d.month, "historical_xlsx"))
    insert_batch(conn, rows)
    print(f"  Joint xlsx: {len(rows)} imported, {skipped} skipped")
    return len(rows)


def import_xlsx_perso(conn):
    wb = openpyxl.load_workbook(HIST_PERSO, data_only=True)
    ws = wb["PERSO DATA"]
    rows = []
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = row[0]
        if not d or not hasattr(d, 'year'):
            skipped += 1
            continue
        date_str = d.strftime("%Y-%m-%d")
        desc = row[1] or ""
        amount = row[2]
        if amount is None:
            skipped += 1
            continue
        category = row[4] or ""   # col 4 = CATEGORY (col 3 = TYPE)
        subcategory = row[5] or ""
        rows.append((date_str, desc, float(amount), "personal",
                     category, subcategory, d.year, d.month, "historical_xlsx"))
    insert_batch(conn, rows)
    print(f"  Personal xlsx: {len(rows)} imported, {skipped} skipped")
    return len(rows)


def import_json(conn, path, account):
    with open(path) as f:
        data = json.load(f)

    # Derive year/month from filename if missing in JSON
    basename = os.path.basename(path)  # e.g. 2026-04-joint.json
    year = data.get("year") or int(basename[:4])
    month = data.get("month") or int(basename[5:7])
    source = f"{year:04d}-{month:02d}-json"
    txns = data.get("transactions", [])

    rows = []
    for t in txns:
        date_str = t.get("date") or f"{year:04d}-{month:02d}-01"
        # Normalize date if it's like "YYYY-MM"
        if date_str and len(date_str) == 7:
            date_str = date_str + "-01"
        desc = t.get("description") or ""
        amount = t.get("amount")
        if amount is None:
            continue
        category = t.get("category") or ""
        subcategory = t.get("subcategory") or ""
        rows.append((date_str, desc, float(amount), account,
                     category, subcategory, year, month, source))

    insert_batch(conn, rows)
    return len(rows)


def import_all_jsons(conn):
    total = 0
    for pattern, account in [
        (f"{HIST_DIR}/????-??-joint.json", "joint"),
        (f"{HIST_DIR}/????-??-personal.json", "personal"),
    ]:
        for path in sorted(glob.glob(pattern)):
            n = import_json(conn, path, account)
            fname = os.path.basename(path)
            print(f"  {fname}: {n} transactions")
            total += n
    return total


def import_income_xlsx(conn):
    HIST_INCOME = "/workspace/extra/historical/historical_data_income.xlsx"
    wb = openpyxl.load_workbook(HIST_INCOME, data_only=True)
    ws = wb["INCOME"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = row[0]
        if not d or not hasattr(d, 'year'):
            continue
        who = row[1] or ""
        value = row[2]
        typ = row[3] or ""
        if value is None:
            continue
        rows.append((d.strftime("%Y-%m-%d"), who, typ, float(value), d.year, d.month, "historical_data_income.xlsx"))
    conn.executemany(
        "INSERT INTO income (date, who, type, amount, year, month, source) VALUES (?,?,?,?,?,?,?)",
        rows
    )
    conn.commit()
    print(f"  Income xlsx: {len(rows)} imported")
    return len(rows)


def import_income_jsons(conn):
    total = 0
    for f in sorted(glob.glob(f"{HIST_DIR}/????-??-income.json")):
        basename = os.path.basename(f)
        m_match = __import__('re').match(r'(\d{4})-(\d{2})-income', basename)
        year, month = int(m_match.group(1)), int(m_match.group(2))
        with open(f) as fp:
            data = json.load(fp)
        rows = []
        for e in data.get("entries", []):
            rows.append((e["date"], e["who"], e["type"], float(e["value"]), year, month, basename))
        conn.executemany(
            "INSERT INTO income (date, who, type, amount, year, month, source) VALUES (?,?,?,?,?,?,?)",
            rows
        )
        conn.commit()
        print(f"  {basename}: {len(rows)} income entries")
        total += len(rows)
    return total


def main():
    print(f"Creating/opening {DB_PATH}")

    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
        print("  Removed existing DB for clean migration")

    conn = connect()
    create_schema(conn)

    print("\nImporting transaction xlsx...")
    n1 = import_xlsx_commun(conn)
    n2 = import_xlsx_perso(conn)

    print("\nImporting transaction JSONs (Apr 2026+)...")
    n3 = import_all_jsons(conn)

    print("\nImporting income data...")
    n4 = import_income_xlsx(conn)
    n5 = import_income_jsons(conn)

    print(f"\nTransactions: {n1+n2+n3}")
    print(f"Income entries: {n4+n5}")

    cur = conn.execute("""
        SELECT account, MIN(date) as earliest, MAX(date) as latest, COUNT(*) as n
        FROM transactions GROUP BY account ORDER BY account
    """)
    print("\nTransaction coverage:")
    for row in cur:
        print(f"  {row['account']}: {row['n']} ({row['earliest']} → {row['latest']})")

    r = conn.execute("SELECT MIN(date), MAX(date), COUNT(*) FROM income").fetchone()
    print(f"\nIncome coverage: {r[2]} entries ({r[0]} → {r[1]})")

    conn.close()
    print(f"\nDone. DB saved to {DB_PATH}")


if __name__ == "__main__":
    main()
