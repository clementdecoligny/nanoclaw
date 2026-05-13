#!/opt/wpenv/bin/python3
"""Write monthly income entries to .xlsx matching the historical income file format.

Output columns (identical to historical_data_income.xlsx):
  Date | Who | Value | Type

Followed by a summary block showing per-person totals and joint contribution split.

Usage:
  income_writer.py --input /tmp/income.json --output /tmp/2026-05-revenus.xlsx
"""

import json
import argparse
from decimal import Decimal, ROUND_HALF_UP
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

VALID_TYPES = {"SALARY", "FOOD", "EDUCATION", "LICENCA", "BONUS", "HOLIDAY", "IRS"}
VALID_WHO = {"Clément", "Lola", "Joint"}

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUMMARY_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
SUMMARY_FONT = Font(bold=True)
WARN_FILL = PatternFill(start_color="FFE0CC", end_color="FFE0CC", fill_type="solid")

HEADERS = ["Date", "Who", "Value", "Type"]


def _cents(v):
    return Decimal(str(v)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def write_income_xlsx(data, output_path):
    entries = data["entries"]
    joint_target = _cents(data.get("joint_target", 4500.0))

    # Sort: date ascending, then who
    entries = sorted(entries, key=lambda e: (e["date"], e["who"]))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "INCOME"

    # Header row
    ws.append(HEADERS)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for e in entries:
        ws.append([e["date"], e["who"], float(_cents(e["value"])), e["type"]])

    # Blank separator
    ws.append([])

    # Compute per-person totals (exclude Joint from proportional calc)
    totals = {"Clément": Decimal("0"), "Lola": Decimal("0"), "Joint": Decimal("0")}
    for e in entries:
        who = e["who"]
        if who in totals:
            totals[who] += _cents(e["value"])

    personal_total = totals["Clément"] + totals["Lola"]
    grand_total = personal_total + totals["Joint"]

    # Proportional joint contributions based on personal income only
    if personal_total > 0:
        clement_share = (_cents(totals["Clément"] / personal_total * joint_target))
        lola_share = joint_target - clement_share
    else:
        clement_share = Decimal("0")
        lola_share = Decimal("0")

    surplus = grand_total - joint_target

    summary_rows = [
        ("Total Clément", float(totals["Clément"])),
        ("Total Lola", float(totals["Lola"])),
        ("Total Joint", float(totals["Joint"])),
        ("Grand total", float(grand_total)),
        ("", ""),
        ("Part compte commun — Clément", float(clement_share)),
        ("Part compte commun — Lola", float(lola_share)),
        ("Objectif compte commun", float(joint_target)),
        ("Excédent / Manque", float(surplus)),
    ]

    shortfall = surplus < 0

    for label, value in summary_rows:
        row_idx = ws.max_row + 1
        ws.cell(row=row_idx, column=1, value=label)
        if value != "":
            ws.cell(row=row_idx, column=3, value=value)
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if label == "Excédent / Manque" and shortfall:
                cell.fill = WARN_FILL
            else:
                cell.fill = SUMMARY_FILL
            cell.font = SUMMARY_FONT

    # Column widths
    widths = [14, 12, 12, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Format Value column as number with 2 decimal places
    for row in ws.iter_rows(min_row=2):
        cell = row[2]  # column C = Value
        if isinstance(cell.value, float):
            cell.number_format = '#,##0.00'

    wb.save(output_path)

    return {
        "clement_total": float(totals["Clément"]),
        "lola_total": float(totals["Lola"]),
        "joint_total": float(totals["Joint"]),
        "grand_total": float(grand_total),
        "clement_share": float(clement_share),
        "lola_share": float(lola_share),
        "joint_target": float(joint_target),
        "surplus": float(surplus),
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="JSON file with income entries")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    args = parser.parse_args()

    with open(args.input, encoding="utf-8") as f:
        data = json.load(f)

    summary = write_income_xlsx(data, args.output)

    # Print summary to stdout for Edmond to use in the chat message
    import json as _json
    print(_json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
