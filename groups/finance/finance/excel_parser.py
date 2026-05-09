#!/opt/wpenv/bin/python3
"""Parse an ActivoBank .xlsx export into JSON.

Rules:
- Data Valor is the only date field used (Data Lanç. is ignored)
- Rows with blank description but non-zero amount are kept as UNCLASSIFIED
- Rows with blank description and zero amount are skipped
- Output is sorted by Data Valor ascending
- Target month filtering is done by the caller (categorizer)
"""

import json
import argparse
import openpyxl
from utils import parse_date, parse_amount, find_description_header_row

ACCOUNT_NUMBERS = {
    "45507717811": "personal",
    "45545535104": "joint",
}


def detect_account(ws):
    """Scan metadata rows for a known account number. Returns 'personal', 'joint', or None."""
    for row in ws.iter_rows(max_row=10, values_only=True):
        for cell in row:
            if cell:
                s = str(cell).replace(" ", "").replace(".", "")
                for number, account in ACCOUNT_NUMBERS.items():
                    if number in s:
                        return account
    return None


def parse_file(file_path, account=None):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    detected = detect_account(ws)
    if detected:
        account = detected

    header_row = find_description_header_row(ws)
    headers = [str(c.value).strip() if c.value else "" for c in ws[header_row]]

    col = {}
    for i, h in enumerate(headers):
        hl = h.lower()
        if "valor" in hl and "data" in hl:
            col["value_date"] = i
        elif "descri" in hl:
            col["description"] = i
        elif hl == "valor":
            col["amount"] = i
        elif "saldo" in hl:
            col["balance"] = i

    transactions = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(v is None for v in row):
            continue

        amount = parse_amount(row[col["amount"]]) if "amount" in col else 0.0
        raw_desc = row[col["description"]] if "description" in col else None
        desc = str(raw_desc).strip() if raw_desc else ""

        # Skip rows with no description and no amount
        if not desc and amount == 0.0:
            continue

        transactions.append({
            "value_date": parse_date(row[col["value_date"]]) if "value_date" in col else None,
            "description": desc if desc else "(sans description)",
            "amount": amount,
            "balance": parse_amount(row[col["balance"]]) if "balance" in col else 0.0,
            "account": account,
            # Pre-flag blank-description rows as UNCLASSIFIED
            "category": "UNCLASSIFIED" if not desc else "",
            "sub_category": "UNCLASSIFIED" if not desc else "",
            "confidence": "forced" if not desc else "",
        })

    # Sort by Data Valor ascending
    transactions.sort(key=lambda t: t["value_date"] or "")

    return transactions, account


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("file_path")
    parser.add_argument("--account", choices=["personal", "joint"], default=None,
                        help="Override account type (auto-detected from account number if omitted)")
    parser.add_argument("--output", default=None)
    args = parser.parse_args()

    transactions, detected_account = parse_file(args.file_path, args.account)
    out = json.dumps({"transactions": transactions, "account": detected_account},
                     ensure_ascii=False, indent=2)

    if args.output:
        with open(args.output, "w", encoding="utf-8") as f:
            f.write(out)
    else:
        print(out)


if __name__ == "__main__":
    main()
