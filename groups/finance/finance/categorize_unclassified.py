#!/opt/wpenv/bin/python3
"""Categorize all UNCLASSIFIED transactions from historical_data_commun.xlsx."""
import sys, re
sys.path.insert(0, '/workspace/agent/finance')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from utils import parse_date, parse_amount

# ─── Categorization rules (order matters — first match wins) ───────────────
RULES = [
    # ── EAT OUT ──
    (r'TGTG|TOO GOOD TO GO', 'GROCERIES', 'SUPER'),          # Too Good To Go app
    (r'GLEBA|GLEBA PRINCIPE|GLEBA NOSSA', 'EAT OUT', 'BACKERY'),
    (r'MANTEIGARIA|A MANTEIGARIA', 'EAT OUT', 'BACKERY'),
    (r'PAO DE CANELA|PADARIA|PANETERIE', 'EAT OUT', 'BACKERY'),
    (r'BUNA ', 'EAT OUT', 'BACKERY'),
    (r'MU GELATO|GELADOS MATILORI|GIOLA GEL|SANTINI|BOZZOLO', 'EAT OUT', 'RESTAURANT'),
    (r'NATA PORTUGUESA|MARIE BLACHERE|ARCE BOUTIQUE DOCE|DOBECO|CAFE DA GEMA', 'EAT OUT', 'BACKERY'),
    (r'VEDETA POSITIVA', 'EAT OUT', 'RESTAURANT'),
    (r'SABOROSO', 'EAT OUT', 'RESTAURANT'),
    (r'NIVA 41', 'EAT OUT', 'RESTAURANT'),
    (r'GLORIA A SECO', 'EAT OUT', 'RESTAURANT'),
    (r'BARTOLO VALE', 'EAT OUT', 'RESTAURANT'),
    (r'WOORI', 'EAT OUT', 'RESTAURANT'),
    (r'PICANHA LISBOA', 'EAT OUT', 'RESTAURANT'),
    (r'BLIND TIGER', 'EAT OUT', 'RESTAURANT'),
    (r'VENTURA.*LICITRA|VENTURA E LICITRA', 'EAT OUT', 'RESTAURANT'),
    (r'MERO MAE', 'EAT OUT', 'RESTAURANT'),
    (r'SUMUP RAIZES DO BRASIL', 'EAT OUT', 'RESTAURANT'),
    (r'RC SANCHES', 'EAT OUT', 'RESTAURANT'),
    (r'SUMUP CRAZY BUTCHER', 'EAT OUT', 'RESTAURANT'),
    (r'PEROLA DE ALCANTARA', 'EAT OUT', 'RESTAURANT'),
    (r'ESPACO ZARCO', 'EAT OUT', 'RESTAURANT'),
    (r'FABLE LISBOA', 'EAT OUT', 'RESTAURANT'),
    (r'MAAT CAFE', 'EAT OUT', 'RESTAURANT'),
    (r'CAFE MIMOSKRIDOS', 'EAT OUT', 'RESTAURANT'),
    (r'EL CANDIE|KIOSKO ATLANTERRA', 'EAT OUT', 'RESTAURANT'),
    (r'LA MORONERA|SF ASADOR|CAFETERIA M\.A\.S\.', 'EAT OUT', 'RESTAURANT'),
    (r'DIAS FLORESCENTES', 'EAT OUT', 'RESTAURANT'),
    (r'REAL PRINCIPE', 'EAT OUT', 'RESTAURANT'),
    (r'THE ALBATROZ ATELIER', 'EAT OUT', 'RESTAURANT'),
    (r'VASCO PINTO LDA', 'EAT OUT', 'RESTAURANT'),
    (r'QUIO GENGIBRE', 'EAT OUT', 'RESTAURANT'),
    (r'MANUEL RUI A NABEIRO|LOJA DELTA Q|ESPRESSO LAB|BUNA\b', 'EAT OUT', 'BACKERY'),
    (r'COPENHAGEN LISBOA', 'EAT OUT', 'BACKERY'),
    (r'SUMUP APEWOOD', 'EAT OUT', 'RESTAURANT'),
    (r'FABIRICA VERDADEIRAS SIN|FABRICA VERDADEIRAS SIN', 'EAT OUT', 'BACKERY'),
    (r'HONORIO BARREIROS', 'EAT OUT', 'RESTAURANT'),
    (r'PAYPAY RIBEIRA BRAVA|PAYPAYUE.*RIBEIRA BRAVA', 'EAT OUT', 'RESTAURANT'),
    (r'ESPLANADA XANA', 'EAT OUT', 'RESTAURANT'),
    (r'DRAMATICO LISBOA', 'EAT OUT', 'RESTAURANT'),  # theater bar
    (r'MAISON.*LUCE', 'EAT OUT', 'RESTAURANT'),

    # ── GROCERIES ──
    (r'CHEN JINGQIANG|MOHAMMAD HARUN|MERCEARIA LUIS|MERCEARIA DA ANITA|MERCEARIA ALBERTO ROSA|MERCEAR ALBERTO ROSA', 'GROCERIES', 'SUPER'),
    (r'SARKAR MINIMERCADO|LOW PRICE LISBOA|MINIMERCADO FRUTARIA|RAFA MINIMER FRUT|MINIPRECO', 'GROCERIES', 'SUPER'),
    (r'FERREIRA DE ALMEIDA AZU', 'GROCERIES', 'SUPER'),
    (r'ECOSEIVA', 'GROCERIES', 'SUPER'),
    (r'CELEIRO DIETA', 'GROCERIES', 'SUPER'),
    (r'NORMALAS|VIDA PLENA', 'GROCERIES', 'SUPER'),
    (r'SUMUP TREEGO SOURDOUGH|TREEGO SOURDOUGH', 'GROCERIES', 'SUPER'),
    (r'MANUEL TAVARES LISBOA', 'GROCERIES', 'SUPER'),
    (r'GARCIA E GIL', 'GROCERIES', 'SUPER'),
    (r'LAMIRE', 'GROCERIES', 'SUPER'),
    (r'MELTEJO ALMADA', 'GROCERIES', 'SUPER'),     # could also be MOBILITY; small amt → EAT OUT
    (r'MY AUCHAN', 'GROCERIES', 'SUPER'),
    (r'QUINTA MELIAS', 'GROCERIES', 'SUPER'),

    # ── MOBILITY ──
    (r'CARRIS|TRANSTEJO|CINTURA LSA TVM|CASCAIS AMA TVM|CLESS TICKET ATM MILANO', 'MOBILITY', 'PUBLIC TRANSPORT'),
    (r'CP PARIS|CP SANTOS', 'MOBILITY', 'PUBLIC TRANSPORT'),
    (r'BOLT\.EU', 'MOBILITY', 'BOLT'),
    (r'INTERPARKING|APARC CONCES ROSA AMARI|PARQUES TEJO', 'VOITURE', 'MISC'),

    # ── VOITURE ──
    (r'EST SERVICO TALAVERA|EST SERVICO AP-4 LOS PALACIOS|EST SERVICO.*VISTA HERMOSA|EST SERVICO E\.S\. VISTA', 'VOITURE', 'ESSENCE'),

    # ── LEISURE ──
    (r'UCI EL CORTE INGLES|UCI EL CORTE', 'LEISURE', 'CONCERT'),
    (r'SHOTGUN ANDCO', 'LEISURE', 'CONCERT'),
    (r'LIVE EXPERIENCES', 'LEISURE', 'MISC'),
    (r'LOJA DAS MAQUETAS', 'LEISURE', 'MISC'),
    (r'BOL QUEIJAS', 'LEISURE', 'MISC'),
    (r'RETROSARIA CHIQUE', 'LEISURE', 'MISC'),   # fabric/craft shop

    # ── EAT OUT (additional) ──
    (r'THE STREET FOOD SPOT|MISS CAN LISBOA|EASY VENDING|LAURINDO LUCILIA', 'EAT OUT', 'RESTAURANT'),
    (r'GLOVO', 'GROCERIES', 'SUPER ONLINE'),
    (r'PAPELARIA FERNANDES|PAPELARIA PLANETA', 'KIDS', 'MISC'),

    # ── KIDS ──
    (r'ESCOLA ALEMA LISBOA', 'EDUCATION', 'ESCOLA'),
    (r'R\. S AVENTUREIRAS|CENTRO SOC PAR N SENHORA|CENTRO SOC PAR', 'KIDS', 'LEISURE'),
    (r'NYX ABSERVICOS|NYX AB', 'KIDS', 'ROPA'),      # NYX cosmetics/accessories
    (r'CONTO FANTASIA', 'KIDS', 'LEISURE'),
    (r'TROPHYCHAMP', 'KIDS', 'LEISURE'),
    (r'QQLI LOJAS', 'KIDS', 'ROPA'),
    (r'ARTCOR', 'KIDS', 'LEISURE'),
    (r'STUFF OUT', 'KIDS', 'ROPA VINTED'),
    (r'LA COSTURA DE MI RATON', 'KIDS', 'ROPA'),

    # ── HEALTH ──
    (r'DR.*CARLOS MIGUEL BORGES|DR.*RODOLFO FREDERICO BEJA', 'HEALTH', 'HOSPITAL'),
    (r'FCIA\. FDEZ DHERBE|FARMACIA|FCIA\.', 'HEALTH', 'PHARMACY'),
    (r'PAGSERV MDGD', 'HEALTH', 'MEDIS REIMBURSMENT'),
    (r'MAISON LUCE', 'HEALTH', 'PHARMACY'),   # health/beauty boutique

    # ── HOUSE ──
    (r'BENOIT BERNARD PHILIPPE DECRETON', 'HOUSE', 'FURNITURE'),
    (r'DROGARIA SAO MARCAL', 'HOUSE', 'MISC'),
    (r'ALTICE PAY', 'HOUSE', 'NOS INTERNET'),
    (r'WEAREODESTUDIOS', 'HOUSE', 'MISC'),
    (r'AFONSO E LOURENCO', 'HOUSE', 'MISC'),
    (r'LOJA S\. BENTO', 'HOUSE', 'FURNITURE'),
    (r'LOJA SENTO|HONORIO BARREIROS E Lisboa', 'HOUSE', 'FURNITURE'),

    # ── HOLIDAY ──
    (r'RYANAIR', 'HOLIDAY', 'TRANSPORT'),

    # ── MISC ──
    (r'CTT.*CORREIOS|CTT - CORREIOS', 'MISC', 'MISC'),
    (r'CUSTO DE SERVICO INTERNACIONAL|IMPOSTO DO SELO', 'MISC', 'MISC'),
    (r'0061320190073338|CHARGES', 'MISC', 'MISC'),
    (r'EUPAGO', 'MISC', 'MISC'),
    (r'HIPAY', 'MISC', 'MISC'),
    (r'HIPAY LEVALLOIS', 'MISC', 'MISC'),
    (r'PAYSHOP', 'MISC', 'MISC'),
    (r'PAYPAY\b|PAYPAYUE', 'MISC', 'MISC'),
    (r'AMAZON', 'MISC', 'MISC'),
    (r'SHCZ SMARTPHONE|SMARTPHONE CLIN', 'MISC', 'MISC'),
    (r'FOTO ESTUDIO', 'MISC', 'MISC'),
    (r'SANTIRIA GESTAO', 'MISC', 'MISC'),
    (r'NELIA SOFIA|GABRIEL GAMBARINI|RICARDO JORGE SALVADO', 'MISC', 'MISC'),
    (r'SUSANA ISABEL', 'MISC', 'MISC'),
    (r'WILD ESTIMATE', 'MISC', 'MISC'),
    (r'MIRIAM SILVA FERNAO', 'MISC', 'MISC'),
    (r'MARIA OLIVEIRA S G GALANTE', 'MISC', 'MISC'),
    (r'MAS GO FELIPE', 'MISC', 'MISC'),
    (r'MUSIC ARTE', 'MISC', 'MISC'),
    (r'GRUPO PORTO EDITORA', 'MISC', 'MISC'),
    (r'HIPAY LEVALLOIS-PERRET', 'MISC', 'MISC'),
    (r'NORMALAS LISBOA', 'GROCERIES', 'SUPER'),
    (r'BOZZOLO', 'EAT OUT', 'RESTAURANT'),
    (r'F FONTES PEREIRA', 'EAT OUT', 'RESTAURANT'),
    (r'ULTIMA ESSENCIA', 'EAT OUT', 'RESTAURANT'),
    (r'TURIM BOULEVARD', 'EAT OUT', 'RESTAURANT'),
    (r'RAFAEL GUZMAN', 'EAT OUT', 'RESTAURANT'),
    (r'MAISA S\.C\.', 'EAT OUT', 'RESTAURANT'),
    (r'ROMANO BEGUM|ROMA BEGUM', 'EAT OUT', 'RESTAURANT'),
    (r'MONZOOR HOSSAIN', 'EAT OUT', 'RESTAURANT'),
    (r'SABOROSO', 'EAT OUT', 'RESTAURANT'),
]

# Transfers to named individuals — try to match babysitters known from data
BABYSITTER_PATTERNS = [
    r'CONSTANCA.*PIRES', r'MATILDE PEREIRA.*CASTRO', r'MARIA JOAO.*ABRANCHES',
    r'MARIA LUCIA VELASCO', r'MARTA MARIA.*GONCALVES', r'ROSA VELOSO',
    r'ANDRE MARIA.*SILVA REBELO', r'CARLOS MANUEL MARTINS GOMES',
    r'RAQUEL.*ROSADO', r'RUTE.*ESPIRITO SANTO', r'RUTE DO ESP',
    r'PEDRO MELO', r'LEONARDO JOSE FERNANDES',
]
HEALTH_TRANSFER = [r'DR.*CARLOS MIGUEL BORGES', r'DR.*RODOLFO FREDERICO BEJA']
GROCERIES_TRANSFER = [r'MARIA LUCIA VELASCO ARBOLI', r'MICHELE MAMEDE',
                      r'M ISABEL GONCALVES', r'GABRIEL GAMBARINI', r'MARTINHO PEDROSA']

def categorize(desc):
    d = desc.upper()
    # Anonymous transfers
    if re.search(r'TRF MB WAY P/ \*{3,}', d):
        return 'UNCLASSIFIED', 'UNCLASSIFIED', 'low'
    # Named health transfers
    for pat in HEALTH_TRANSFER:
        if re.search(pat, d, re.I):
            return 'HEALTH', 'HOSPITAL', 'high'
    # Named grocery-style transfers (shared expenses)
    for pat in GROCERIES_TRANSFER:
        if re.search(pat, d, re.I):
            return 'GROCERIES', 'SUPER', 'medium'
    # Babysitter transfers
    for pat in BABYSITTER_PATTERNS:
        if re.search(pat, d, re.I):
            return 'BABYSITTING', 'BABYSITTING', 'medium'
    # Rule-based
    for pattern, cat, sub in RULES:
        if re.search(pattern, d, re.I):
            return cat, sub, 'high'
    return 'UNCLASSIFIED', 'UNCLASSIFIED', 'low'

# Load and process
wb_in = openpyxl.load_workbook('/workspace/extra/historical/historical_data_commun.xlsx', data_only=True)
ws_in = wb_in['COMUN DATA']
headers = [str(c.value).strip() if c.value else '' for c in ws_in[1]]
col = {}
for i, h in enumerate(headers):
    hu = h.upper()
    if 'DATE' in hu: col['date'] = i
    elif 'DESCRI' in hu or 'DESCRIPTION' in hu: col['desc'] = i
    elif hu == 'CATEGORY': col['cat'] = i
    elif hu == 'SUB-CATEGORY': col['sub'] = i
    elif hu == 'VALOR': col['amount'] = i

rows = []
for row_num in range(2, ws_in.max_row + 1):
    desc_val = ws_in.cell(row=row_num, column=col['desc']+1).value
    cat_val  = ws_in.cell(row=row_num, column=col['cat']+1).value
    date_val = ws_in.cell(row=row_num, column=col['date']+1).value
    amt_val  = ws_in.cell(row=row_num, column=col['amount']+1).value
    if not desc_val: continue
    cat = str(cat_val).strip().upper() if cat_val else ''
    if not cat or cat == 'UNCLASSIFIED':
        date = parse_date(date_val) if date_val else ''
        amt  = parse_amount(amt_val) if amt_val is not None else 0.0
        new_cat, new_sub, conf = categorize(str(desc_val).strip())
        rows.append({
            'row': row_num, 'date': date,
            'description': str(desc_val).strip(),
            'amount': amt,
            'category': new_cat, 'sub_category': new_sub, 'confidence': conf,
        })

# Build output xlsx
HEADER_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT  = Font(color="FFFFFF", bold=True)
HIGH_FILL    = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
MED_FILL     = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
LOW_FILL     = PatternFill(start_color="FFE0CC", end_color="FFE0CC", fill_type="solid")

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "UNCLASSIFIED propositions"

hdrs = ['Date', 'Description', 'Montant', 'CATEGORIE proposée', 'SOUS-CATEGORIE', 'Confiance']
for j, h in enumerate(hdrs, 1):
    c = ws_out.cell(row=1, column=j, value=h)
    c.fill = HEADER_FILL; c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center')

conf_counts = {'high': 0, 'medium': 0, 'low': 0}
for i, r in enumerate(rows, 2):
    fill = HIGH_FILL if r['confidence'] == 'high' else MED_FILL if r['confidence'] == 'medium' else LOW_FILL
    vals = [r['date'], r['description'], r['amount'], r['category'], r['sub_category'], r['confidence']]
    for j, v in enumerate(vals, 1):
        c = ws_out.cell(row=i, column=j, value=v)
        c.fill = fill
        if j == 3: c.number_format = '#,##0.00'
    conf_counts[r['confidence']] += 1

widths = [12, 55, 10, 20, 20, 12]
for j, w in enumerate(widths, 1):
    ws_out.column_dimensions[get_column_letter(j)].width = w

ws_out.freeze_panes = 'A2'

out_path = '/workspace/agent/finance/unclassified-propositions.xlsx'
wb_out.save(out_path)

print(f'Saved: {out_path}')
print(f'Total: {len(rows)} rows')
print(f'  High confidence:   {conf_counts["high"]}')
print(f'  Medium confidence: {conf_counts["medium"]}')
print(f'  Low / still UNCL: {conf_counts["low"]}')
print()
print('Still UNCLASSIFIED (low confidence):')
for r in rows:
    if r['confidence'] == 'low':
        print(f"  {r['date']}  {r['description'][:55]:<55}  {r['amount']:8.2f}")
