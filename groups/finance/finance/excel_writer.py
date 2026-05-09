#!/opt/wpenv/bin/python3
"""Write categorized transactions to .xlsx matching the historical data format.

Output columns (identical to historical files so output can feed back into history):
  Date | Description | Valor | CATEGORY | SUB-CATEGORY

Rows are sorted by Date (Data Valor) ascending.
UNCLASSIFIED rows are highlighted in orange.
"""

import json
import argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CAT_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
UNCLASSIFIED_FILL = PatternFill(start_color="FFE0CC", end_color="FFE0CC", fill_type="solid")

# Matches historical file format exactly
HEADERS = ["Date", "Description", "Valor", "CATEGORY", "SUB-CATEGORY"]
CAT_COLS = {"CATEGORY", "SUB-CATEGORY"}


def output_filename(year, month, account):
    """e.g. 2026-05-personnel-categorise.xlsx"""
    label = "personnel" if account == "personal" else "commun"
    return f"{year:04d}-{month:02d}-{label}-categorise.xlsx"


def write_xlsx(transactions, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    ws.append(HEADERS)
    cat_col_indices = [i + 1 for i, h in enumerate(HEADERS) if h in CAT_COLS]

    for col_idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for t in transactions:
        ws.append([
            t.get("value_date", ""),
            t.get("description", ""),
            t.get("amount", 0.0),
            t.get("category", "UNCLASSIFIED"),
            t.get("sub_category", ""),
        ])

        row_idx = ws.max_row
        if t.get("category", "UNCLASSIFIED") in ("UNCLASSIFIED", ""):
            for col_idx in range(1, len(HEADERS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = UNCLASSIFIED_FILL
        else:
            for col_idx in cat_col_indices:
                ws.cell(row=row_idx, column=col_idx).fill = CAT_FILL

    widths = [12, 50, 10, 20, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("transactions_json", help="Categorizer output JSON")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    args = parser.parse_args()

    with open(args.transactions_json, encoding="utf-8") as f:
        data = json.load(f)

    transactions = data if isinstance(data, list) else data.get("transactions", [])

    # Ensure sorted by Data Valor ascending
    transactions.sort(key=lambda t: t.get("value_date") or "")

    write_xlsx(transactions, args.output)


if __name__ == "__main__":
    main()
