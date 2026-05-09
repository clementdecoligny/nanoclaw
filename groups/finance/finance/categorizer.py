#!/opt/wpenv/bin/python3
"""
Categorize transactions using a two-pass hybrid strategy:
  Pass 1 — exact lookup against historical data (fast, free, deterministic)
  Pass 2 — emit uncategorized rows as JSON for Edmond to classify via Claude
"""

import sys
import json
import argparse
from pathlib import Path
import openpyxl
from utils import normalize, find_category_header_row


TAXONOMY = {
    "BABYSITTING": ["BABYSITTING"],
    "EAT OUT": ["BACKERY", "PADARIA", "QUIOSQUE", "RESTAURANT", "UBER EATS"],
    "EDUCATION": ["ESCOLA"],
    "EMPREGADA": ["EMPREGADA", "SEGURANCA SOCIAL"],
    "GROCERIES": ["SUPER", "SUPER ONLINE"],
    "HEALTH": ["COUCHES", "HOSPITAL", "MEDIS", "MEDIS REIMBURSMENT", "PHARMACY"],
    "HOLIDAY": ["ALPS JULY 2026", "HOTELS", "MISC", "TRANSPORT"],
    "HOUSE": ["DONA AJUDA", "EDP", "EPAL AGUA", "FURNITURE", "MISC", "NOS", "NOS INTERNET", "SPOTIFY"],
    "KIDS": ["BABYSITTING", "LEISURE", "MISC", "PISCINA", "ROPA", "ROPA VINTED"],
    "MOBILITY": ["BOLT", "COOLTRA", "PUBLIC TRANSPORT", "TRANSPORT", "UBER"],
    "RENT": ["RENT"],
    "VOITURE": ["ESSENCE", "MECHANIC", "VIA VERDE"],
    "CASH": ["CASH"],
    "LEISURE": [],
    "FURNITURE": ["FURNITURE"],
    "GIFT": ["GIFT"],
    "MISC": ["MISC"],
    "UNCLASSIFIED": ["UNCLASSIFIED"],
}


def _parse_history_xlsx(xlsx_path):
    """Extract {normalized_description: {category, sub_category}} from one history file."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    header_row_idx = find_category_header_row(ws)
    if not header_row_idx:
        return {}

    col = {}
    for i, cell in enumerate(ws[header_row_idx]):
        if not cell.value:
            continue
        h = str(cell.value).strip().upper()
        if h in ("DESCRIPTION", "DESCRIÇÃO"):
            col["desc"] = i
        elif h == "CATEGORY":
            col["cat"] = i
        elif h == "SUB-CATEGORY":
            col["sub"] = i

    if "desc" not in col or "cat" not in col:
        return {}

    entries = {}
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        desc = row[col["desc"]] if len(row) > col["desc"] else None
        cat = row[col["cat"]] if len(row) > col["cat"] else None
        sub = row[col["sub"]] if "sub" in col and len(row) > col["sub"] else None
        if desc and cat:
            key = normalize(str(desc))
            if key not in entries:
                entries[key] = {
                    "category": str(cat).strip().upper(),
                    "sub_category": str(sub).strip().upper() if sub else "",
                }
    return entries


def load_history(history_dir):
    """Build lookup: normalized_description → {category, sub_category}.

    Reads all .xlsx files (oldest-first so first occurrence wins), then
    overlays learned_categories.json (Claude-confirmed, higher priority).
    Keys are pre-normalized so callers pay normalize() once per transaction,
    not once per lookup.
    """
    lookup = {}
    history_path = Path(history_dir)

    if not history_path.exists():
        return lookup

    for xlsx in sorted(history_path.glob("*.xlsx")):
        try:
            entries = _parse_history_xlsx(xlsx)
            for key, val in entries.items():
                if key not in lookup:
                    lookup[key] = val
        except Exception as e:
            print(f"Warning: could not read {xlsx}: {e}", file=sys.stderr)

    learned_path = history_path / "learned_categories.json"
    if learned_path.exists():
        try:
            with open(learned_path, encoding="utf-8") as f:
                learned = json.load(f)
            # learned keys may be raw descriptions — normalize them on load
            for raw_key, val in learned.items():
                lookup[normalize(raw_key)] = val
        except Exception as e:
            print(f"Warning: could not read learned_categories.json: {e}", file=sys.stderr)

    return lookup


def categorize(transactions, lookup):
    """Apply lookup. Returns (transactions, unknown_indices)."""
    unknown = []
    for i, t in enumerate(transactions):
        key = normalize(t["description"])
        if key in lookup:
            t["category"] = lookup[key]["category"]
            t["sub_category"] = lookup[key]["sub_category"]
            t["confidence"] = "exact"
        else:
            unknown.append(i)
    return transactions, unknown


def build_claude_prompt(transactions, unknown_indices, lookup, sample_size=25):
    """Build prompt payload for Claude to classify unknown transactions."""
    examples = []
    seen_cats = set()
    for key, val in lookup.items():
        cat = val["category"]
        if cat not in seen_cats and len(examples) < sample_size:
            examples.append({"description": key, "category": cat, "sub_category": val["sub_category"]})
            seen_cats.add(cat)

    rows_to_classify = [
        {"index": i, "description": transactions[i]["description"], "amount": transactions[i]["amount"]}
        for i in unknown_indices
    ]

    return {"taxonomy": TAXONOMY, "examples": examples, "rows": rows_to_classify}


def save_learned(history_dir, new_categorizations):
    """Persist new Claude-assigned categorizations to learned_categories.json."""
    learned_path = Path(history_dir) / "learned_categories.json"
    existing = {}
    if learned_path.exists():
        try:
            with open(learned_path, encoding="utf-8") as f:
                existing = json.load(f)
        except Exception:
            pass
    existing.update(new_categorizations)
    with open(learned_path, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("transactions_json")
    parser.add_argument("--history", required=True, help="Directory with historical .xlsx files")
    parser.add_argument("--output", required=True, help="Output JSON path")
    parser.add_argument("--apply-claude", default=None,
                        help="JSON file with Claude's categorization results to apply and persist")
    args = parser.parse_args()

    with open(args.transactions_json, encoding="utf-8") as f:
        transactions = json.load(f)

    lookup = load_history(args.history)
    transactions, unknown_indices = categorize(transactions, lookup)

    if args.apply_claude and Path(args.apply_claude).exists():
        with open(args.apply_claude, encoding="utf-8") as f:
            claude_results = json.load(f)
        new_learned = {}
        for r in claude_results:
            idx = r["index"]
            if 0 <= idx < len(transactions):
                transactions[idx]["category"] = r.get("category", "UNCLASSIFIED")
                transactions[idx]["sub_category"] = r.get("sub_category", "")
                transactions[idx]["confidence"] = "claude"
                new_learned[transactions[idx]["description"]] = {
                    "category": transactions[idx]["category"],
                    "sub_category": transactions[idx]["sub_category"],
                }
        if new_learned:
            save_learned(args.history, new_learned)
        unknown_indices = [i for i in unknown_indices if transactions[i]["confidence"] != "claude"]

    result = {
        "transactions": transactions,
        "unknown_count": len(unknown_indices),
        "exact_count": sum(1 for t in transactions if t.get("confidence") == "exact"),
    }

    if unknown_indices:
        result["claude_prompt"] = build_claude_prompt(transactions, unknown_indices, lookup)
        result["unknown_indices"] = unknown_indices

    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(
        f"Categorized {len(transactions)} transactions: "
        f"{result['exact_count']} exact, {len(unknown_indices)} need Claude.",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
